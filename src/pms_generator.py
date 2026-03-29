"""
Module 9: Final PMS Sheet Generator
Consolidates all data into a complete Piping Material Specification document.
Outputs to Excel and console.
Reference template: pms-A1LN.xlsx (20 NPS sizes, cols B-U, instructions in W-X)
"""

import io
import os
import sys
import tempfile
import urllib.request
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

LOGO_URL = "https://www.shapoorjipallonjienergy.com/img/logo.png"

from data.reference_data import (
    FLANGE_RATINGS_CS, FLANGE_RATINGS_SS,
    FLANGE_RATINGS_CS_METRIC, FLANGE_RATINGS_SS_METRIC,
    FLANGE_RATINGS_CS_GALV_METRIC,
    PIPE_DIMENSIONS_B36_10, PIPE_DIMENSIONS_B36_19,
    STANDARD_OD_MM, VALVE_MATERIALS, VALVE_END_CONNECTIONS,
    VDS_VALVE_TYPE_CODES, VDS_DESIGN_CODES, VDS_END_CONN_CODES,
    VDS_SEAT_SELECTION, VALVE_SIZE_APPLICABILITY,
    get_branch_chart, BRANCH_CHARTS,
)
from src.valve_module import generate_vds_code, recommend_vds_for_class

# 22 NPS sizes matching reference template (cols B-W)
STANDARD_NPS_SIZES = [
    "0.5", "0.75", "1", "1.5", "2", "3", "4", "6", "8", "10",
    "12", "14", "16", "18", "20", "22", "24", "26", "28", "30",
    "32", "36",
]

# ── Pipe Type Selection (per ASME B31.3 & PMS practice) ──────────
# Rules:
#   NPS ≤ 16"  → "Seamless"
#   NPS ≥ 18"  → "LSAW" (default), upgrade to "LSAW, 100% RT" for:
#                 - NACE / sour / critical service
#                 - Hydrocarbon / Gas / Hazardous service
SEAMLESS_SPLIT_NPS = 20  # NPS at which welded pipe starts (Seamless ≤ 18", LSAW ≥ 20")

_HC_KEYWORDS = ["hydrocarbon", "hc ", "gas", "hazardous", "flammable",
                "fuel", "diesel", "crude", "naphtha", "lpg", "lng"]
_SOUR_KEYWORDS = ["sour", "h2s", "sulfide", "nace"]


def determine_pipe_type(nps_str, service="", is_nace=False, is_critical=False):
    """Determine pipe type per ASME B31.3 & industry PMS rules.
    Returns one of: 'Seamless', 'LSAW', 'LSAW, 100% RT'
    """
    nps_f = float(nps_str)
    svc = (service or "").lower()

    # Rule 1: Size-based — Seamless up to NPS 16"
    if nps_f <= 16:
        return "Seamless"

    # NPS >= 20" default is LSAW
    pipe_type = "LSAW"

    # Rule 2: Service override — HC / Gas / Hazardous → LSAW, 100% RT
    if any(k in svc for k in _HC_KEYWORDS):
        pipe_type = "LSAW, 100% RT"

    # Rule 3: NACE / Sour / Critical override → LSAW, 100% RT
    if is_nace or is_critical:
        pipe_type = "LSAW, 100% RT"
    if any(k in svc for k in _SOUR_KEYWORDS):
        pipe_type = "LSAW, 100% RT"

    return pipe_type

# ── Material mappings for NACE / Low Temp / Standard service ──────

# Pipe MOC (seamless) by material type
PIPE_MOC_SEAMLESS = {
    "CS":       "ASTM A 106 Gr. B",
    "CS GALV":  "ASTM A 106 Gr. B",
    "CS_GALV":  "ASTM A 106 Gr. B",
    "CS-LT":    "ASTM A 333 Gr.6",
    "SS":       "ASTM A 312 TP316",
    "Alloy":    "ASTM A 335 P11",
}

# Pipe MOC (welded / LSAW) by material type
PIPE_MOC_WELDED = {
    "CS":       "API 5L Gr. B",
    "CS GALV":  "API 5L Gr. B",
    "CS_GALV":  "API 5L Gr. B",
    "CS-LT":    "ASTM A 671 - CC60 Class 22",
    "SS":       "ASTM A 358 TP316 Class 1",
    "Alloy":    "ASTM A 691 1-1/4 Cr",
}

# Fitting MOC by material type
FITTING_MOC = {
    "CS":       "ASTM A 234 Gr. WPB",
    "CS GALV":  "ASTM A 234 Gr. WPB, Seamless Galvanized",
    "CS_GALV":  "ASTM A 234 Gr. WPB, Seamless Galvanized",
    "CS-LT":    "ASTM A 420 Gr. WPL6",
    "SS":       "ASTM A 403 Gr. WP316",
    "Alloy":    "ASTM A 234 WP11",
}

# Flange / Spectacle MOC by material type
FLANGE_MOC = {
    "CS":       "ASTM A 105",
    "CS GALV":  "ASTM A 105N Galvanized",
    "CS_GALV":  "ASTM A 105N Galvanized",
    "CS-LT":    "ASTM A 350 Gr. LF2",
    "SS":       "ASTM A 182 F316/F316L",
    "Alloy":    "ASTM A 182 F11",
}

# Weldolet / Olet MOC by material type
WELDOLET_MOC = {
    "CS":       "MSS SP 97, ASTM A 105",
    "CS GALV":  "MSS SP 97, ASTM A 105N, Galvanized",
    "CS_GALV":  "MSS SP 97, ASTM A 105N, Galvanized",
    "CS-LT":    "MSS SP 97, ASTM A 350 Gr. LF2",
    "SS":       "MSS SP 97, ASTM A 182 F316",
    "Alloy":    "MSS SP 97, ASTM A 182 F11",
}

# Bolting by NACE flag
_COAT = ", XYLAR 2 + XYLAN 1070 coated with minimum combined thickness of 50\u03bcm"
BOLT_MATERIAL = {
    # (effective_material, is_nace) → stud bolt spec
    ("CS",    False): "ASTM A 193 Gr. B7M" + _COAT,
    ("CS",    True):  "ASTM A 320 Gr. L7M" + _COAT,
    ("CS-LT", False): "ASTM A 320 Gr. L7",
    ("CS-LT", True):  "ASTM A 320 Gr. L7M" + _COAT,
    ("SS",    False): "ASTM A 193 Gr. B8M Class 2",
    ("SS",    True):  "ASTM A 193 Gr. B8M Class 2",
    ("Alloy", False): "ASTM A 193 Gr. B7",
    ("Alloy", True):  "ASTM A 320 Gr. L7M" + _COAT,
}
NUT_MATERIAL = {
    # (effective_material, is_nace) → hex nut spec
    ("CS",    False): "ASTM A 194 Gr. 2HM" + _COAT,
    ("CS",    True):  "ASTM A 194 Gr. 7ML" + _COAT,
    ("CS-LT", False): "ASTM A 194 Gr. 7",
    ("CS-LT", True):  "ASTM A 194 Gr. 7ML" + _COAT,
    ("SS",    False): "ASTM A 194 Gr. 8M",
    ("SS",    True):  "ASTM A 194 Gr. 8M",
    ("Alloy", False): "ASTM A 194 Gr. 2H",
    ("Alloy", True):  "ASTM A 194 Gr. 7ML" + _COAT,
}

# Material display name
MATERIAL_DISPLAY = {
    "CS":       "Carbon Steel",
    "CS-LT":    "LTCS",
    "CS GALV":  "CS Galvanised",
    "CS_GALV":  "CS Galvanised",
    "SS":       "Stainless Steel",
    "DSS":      "Duplex Stainless Steel",
    "SDSS":     "Super Duplex Stainless Steel",
    "CuNi":     "90/10 Copper Nickel",
    "GRE":      "Glass Reinforced Epoxy",
    "Alloy":    "Alloy Steel",
}

# Design code string by NACE flag
DESIGN_CODE = {
    True:  "ASME B 31.3, NACE-MR-01-75/ISO-15156-1/2/3",
    False: "ASME B 31.3",
}

# Gasket description by face type
GASKET_DESC = {
    "RF":  "ASME B 16.20, 4.5mm, SS316/SS316L Spiral Wound with Flexible Graphite (F.G.) filler",
    "RTJ": "ASME B 16.20, Ring Joint, SS316/SS316L",
}


def get_pt_rating_pairs(material_type, flange_class):
    """Extract P-T rating pairs in metric: (temp_c, pressure_bar).
    Uses ASME B16.5-2020 METRIC tables directly (°C and bar).
    Routes CS GALV to the de-rated Group 2.1 table.
    """
    # Select the correct metric table based on material type
    mt = str(material_type or "CS").upper().replace(" ", "_")
    if mt == "SS" or "316" in mt or "304" in mt:
        metric_table = FLANGE_RATINGS_SS_METRIC
    elif mt in ("CS_GALV", "CSGALV", "CS-GALV"):
        metric_table = FLANGE_RATINGS_CS_GALV_METRIC
    else:
        metric_table = FLANGE_RATINGS_CS_METRIC

    flange_class_int = int(flange_class) if flange_class else 150

    pairs = []
    for temp_c in sorted(metric_table.keys()):
        if flange_class_int in metric_table[temp_c]:
            p_bar = metric_table[temp_c][flange_class_int]
            pairs.append((temp_c, p_bar))

    # Fallback to imperial table with conversion if metric not available
    if not pairs:
        ratings = FLANGE_RATINGS_SS if "SS" in mt else FLANGE_RATINGS_CS
        for temp_f in sorted(ratings.keys()):
            if flange_class_int in ratings[temp_f]:
                temp_c = round((temp_f - 32) * 5 / 9)
                p_bar = round(ratings[temp_f][flange_class_int] * 0.0689476, 1)
                pairs.append((temp_c, p_bar))
    return pairs


def _get_face_str(piping_class, face_type):
    """Get face type abbreviation based on class. RF for ≤600#, RTJ for ≥900#."""
    pc = int(piping_class) if piping_class else 150
    if pc >= 900:
        return "RTJ"
    # Normalize face_type to short form
    ft = face_type or "RF"
    if "RF" in ft or "Raised" in ft:
        return "RF"
    if "RTJ" in ft or "Ring" in ft:
        return "RTJ"
    if "FF" in ft or "Flat" in ft:
        return "FF"
    return ft


def _get_valve_codes(pms_code, piping_class, face_str):
    """Generate valve VDS tag codes based on PMS code."""
    code = pms_code or "A1"
    # R = RF (Raised Face), J = RTJ (Ring Type Joint)
    suffix = "J" if face_str == "RTJ" else "R"
    return {
        "Ball":  f"BSRA{code[1:]}{suffix}, BSFA{code[1:]}{suffix}" if len(code) > 1 else "",
        "Gate":  f"GAWA{code[1:]}{suffix}" if len(code) > 1 else "",
        "Globe": f"GLSA{code[1:]}{suffix}" if len(code) > 1 else "",
        "Check": f"CHPA{code[1:]}{suffix}",
        "CheckExtra": f"CSWA{code[1:]}{suffix}, CDPA{code[1:]}{suffix}" if len(code) > 1 else "",
    }


def generate_pms_console(pms_data):
    """Print the complete PMS to console (brief summary)."""
    sc = pms_data.get("spec_code", {})
    msr = pms_data.get("msr", {})
    meta = pms_data.get("metadata", {})
    print(f"\n  PMS: {sc.get('spec_code', 'N/A')} | Material: {msr.get('material_type', 'N/A')} "
          f"| Grade: {msr.get('material_grade', 'N/A')} | Rev: {meta.get('revision', 'C0')}")


def generate_pms_excel(pms_data, output_dir):
    """Generate PMS Excel matching reference template format.

    Layout: 21 data cols (A-U), instruction cols (W-X), 20 NPS sizes (B-U).
    Matches pms-A1LN.xlsx reference template exactly.
    """
    if not HAS_OPENPYXL:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Piping Material Specification"

    # ── Styles ─────────────────────────────────────────────────────
    font_n = Font(name="Aptos Narrow", size=11)
    font_b = Font(name="Aptos Narrow", size=11, bold=True)
    font_instr = Font(name="Aptos Narrow", size=10, italic=True, color="0000FF")
    font_ref = Font(name="Aptos Narrow", size=10, italic=True, color="008000")
    ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ts = Side(style="thin")
    ms = Side(style="medium")
    tb = Border(top=ts, bottom=ts, left=ts, right=ts)
    fill_section = PatternFill("solid", fgColor="D9E1F2")  # Light blue section headers
    fill_input = PatternFill("solid", fgColor="FFF2CC")    # Yellow for user inputs

    LAST_COL = 23  # Column W (22 NPS sizes + 1 label col: A=label, B-W=sizes)
    INSTR_COL = 25  # Column Y (reference notes)
    SOURCE_COL = 26  # Column Z (source labels)

    for col_idx in range(1, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13.0
    ws.column_dimensions[get_column_letter(INSTR_COL)].width = 40.0
    ws.column_dimensions[get_column_letter(SOURCE_COL)].width = 18.0

    def wc(row, col, value, bold=False, left=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font_b if bold else font_n
        c.alignment = al if left else ac
        c.border = tb
        return c

    def mw(row, sc, ec, value, bold=False, left=False):
        if sc < ec:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        c = ws.cell(row=row, column=sc, value=value)
        c.font = font_b if bold else font_n
        c.alignment = al if left else ac
        for col in range(sc, ec + 1):
            ws.cell(row=row, column=col).border = tb
        return c

    def instr(row, text, col=INSTR_COL):
        """Write reference note in column W."""
        c = ws.cell(row=row, column=col, value=text)
        c.font = font_ref
        c.alignment = al

    def source(row, text):
        """Write source label in column X."""
        c = ws.cell(row=row, column=SOURCE_COL, value=text)
        c.font = font_instr
        c.alignment = al

    # ── Extract data ───────────────────────────────────────────────
    meta = pms_data.get("metadata", {})
    msr = pms_data.get("msr", {})
    sc = pms_data.get("spec_code", {})
    line = pms_data.get("line_list", {})
    thick = pms_data.get("thickness", {})
    sched = pms_data.get("schedule", {})
    fittings = pms_data.get("fittings", {})
    flanges_data = pms_data.get("flanges", {})
    fl = flanges_data.get("flange", {})
    gk = flanges_data.get("gasket", {})
    bl = flanges_data.get("bolting", {})
    valve_data = pms_data.get("valves", {})
    valves = valve_data.get("valves", {})

    spec_code = str(sc.get("spec_code", "") or "")
    pms_code  = str(sc.get("pms_code", spec_code) or spec_code)
    # pms_material_type = original user selection (e.g. "CS GALV", "CuNi", "GRE")
    # material_type     = base type for pipe dimension tables (e.g. "CS", "SS")
    pms_material_type = str(msr.get("pms_material_type", "") or msr.get("material_type", "CS") or "CS")
    material_type = str(msr.get("material_type", "CS") or "CS")
    # Coerce to plain bool — JSON may deliver {} / null / 0 / "true" from some clients
    is_nace    = bool(sc.get("is_nace", False)) or "N" in (str(sc.get("part3_code", "") or ""))
    is_low_temp = bool(sc.get("is_low_temp", False)) or "L" in (str(sc.get("part3_code", "") or ""))

    # Determine effective material type for lookups
    eff_mat = str(pms_material_type)
    if material_type == "CS" and is_low_temp:
        eff_mat = "CS-LT"

    part1_info = sc.get("part1_info", {})
    # Authoritative order: spec code Part 1 → spec code pressure_class field → flange class → default 150
    piping_class = (
        part1_info.get("pressure_psig")         # e.g. 150 for Part1="A"
        or sc.get("pressure_class")              # explicit field if stored
        or fl.get("class")                       # auto-selected flange class (fallback)
        or 150                                   # ultimate default
    )
    flange_class = piping_class  # Always align with piping class from spec code
    face_type = _get_face_str(piping_class, fl.get("face_type", "RF"))
    ca_mm = msr.get("corrosion_allowance_mm", 0)

    # Material display name (include NACE suffix if applicable)
    mat_display = MATERIAL_DISPLAY.get(eff_mat, material_type)
    if is_nace and "NACE" not in mat_display:
        mat_display += " NACE"

    # ── Build pipe data from schedule_rows ─────────────────────────
    schedule_rows = sched.get("schedule_rows", [])
    schedule_by_nps = {str(r["nps"]): r for r in schedule_rows} if schedule_rows else {}

    # Service info for pipe type determination
    service_desc = (
        pms_data.get("service", {}).get("service_description", "")
        or line.get("fluid", "")
        or msr.get("fluid_service", "")
    )
    is_critical = bool(pms_data.get("service", {}).get("is_critical", False))

    pipe_sizes = []
    dim_table = PIPE_DIMENSIONS_B36_19 if material_type == "SS" else PIPE_DIMENSIONS_B36_10
    for nps in STANDARD_NPS_SIZES:
        # Use exact metric OD from standard table (avoids rounding errors from in→mm)
        od_mm = STANDARD_OD_MM.get(nps, "-")
        pipe_type = determine_pipe_type(nps, service=service_desc, is_nace=is_nace, is_critical=is_critical)
        if nps in schedule_by_nps:
            r = schedule_by_nps[nps]
            pipe_sizes.append({"nps": nps, "od_mm": od_mm, "schedule": r.get("schedule", "-"), "wt_mm": r.get("wt_mm", "-"), "pipe_type": pipe_type})
        elif nps in dim_table:
            pipe_sizes.append({"nps": nps, "od_mm": od_mm, "schedule": "-", "wt_mm": "-", "pipe_type": pipe_type})
        else:
            pipe_sizes.append({"nps": nps, "od_mm": od_mm, "schedule": "-", "wt_mm": "-", "pipe_type": pipe_type})

    SC = 2  # size start column (B)

    # Find split column index for seamless/welded (NPS 18 = index 13 in list → col 15)
    split_idx = next((i for i, nps in enumerate(STANDARD_NPS_SIZES) if float(nps) >= SEAMLESS_SPLIT_NPS), len(STANDARD_NPS_SIZES))
    split_col = SC + split_idx  # Column where welded starts

    # ══════════════════════════════════════════════════════════════
    # HEADER (Rows 1-6)
    # ══════════════════════════════════════════════════════════════

    # A1:E6 merged block — logo area
    ws.merge_cells(start_row=1, start_column=1, end_row=6, end_column=5)
    for r in range(1, 7):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = tb
            cell.alignment = ac
    # Increase row heights to accommodate larger logo
    for r in range(1, 7):
        ws.row_dimensions[r].height = 20

    # Embed company logo in A1:E6
    try:
        import ssl
        from PIL import Image as PILImage
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        with urllib.request.urlopen(LOGO_URL, timeout=8, context=ctx) as resp:
            logo_bytes = resp.read()
        # Resize to fill A1:D6 (4 cols × 13 width ≈ 330px wide, 6 rows × 18pt ≈ 108px tall)
        pil_img = PILImage.open(io.BytesIO(logo_bytes)).convert("RGBA")
        pil_img = pil_img.resize((330, 108), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil_img.save(buf, format='PNG')
        buf.seek(0)
        img = XLImage(buf)
        ws.add_image(img, 'A1')
    except Exception:
        pass  # logo unavailable — leave merged area blank

    # Row 1: Title
    mw(1, 6, 17, "PIPING MATERIAL SPECIFICATION", bold=True)
    mw(1, 19, LAST_COL, f"Rev: {meta.get('revision', 'C0')}")

    # Row 2: Labels
    mw(2, 6, 9, "Piping Class")
    mw(2, 10, 12, "Material")
    mw(2, 13, 15, "C.A")
    mw(2, 16, 18, "Mill Tol")
    mw(2, 19, LAST_COL, "Sheet No.")

    # Row 3: Values
    mw(3, 6, 7, spec_code)
    wc(3, 8, f"{piping_class}#")
    wc(3, 9, "")
    mw(3, 10, 12, mat_display)
    mw(3, 13, 15, f"{ca_mm} mm")
    mw(3, 16, 18, "12.5%")
    mw(3, 19, LAST_COL, meta.get("revision", "0"))

    # Row 4: Design Code
    mw(4, 6, 8, "Design Code: ", bold=True, left=True)
    mw(4, 9, LAST_COL, DESIGN_CODE.get(bool(is_nace), "ASME B 31.3"))

    # Row 5: Service
    mw(5, 6, 8, "Service:", bold=True, left=True)
    mw(5, 9, LAST_COL, line.get("fluid", msr.get("fluid_service", "General Service")))

    # Row 6: Branch Chart — select based on pms_material_type
    branch_chart_num = get_branch_chart(pms_material_type)
    branch_chart_info = BRANCH_CHARTS.get(branch_chart_num, {})
    branch_chart_title = branch_chart_info.get("title", f"Chart {branch_chart_num}")
    mw(6, 6, 8, "Branch Chart: ", bold=True, left=True)
    mw(6, 9, LAST_COL, f"Ref. APPENDIX-1, {branch_chart_title}")

    # ══════════════════════════════════════════════════════════════
    # PRESSURE-TEMPERATURE RATING (Rows 7-9)
    # ══════════════════════════════════════════════════════════════
    mw(7, 1, LAST_COL, "Pressure-Temperature Rating", bold=True)
    ws.cell(row=7, column=1).fill = fill_section
    for c in range(2, LAST_COL + 1):
        ws.cell(row=7, column=c).fill = fill_section

    pt_pairs = get_pt_rating_pairs(pms_material_type, piping_class)

    # Determine minimum design temperature for material
    # LTCS (A333 Gr.6): -45°C, CS (A106 Gr.B): -29°C, SS: -196°C
    min_temp_c = -29
    if is_low_temp and eff_mat == "CS-LT":
        min_temp_c = -45
    elif material_type == "SS":
        min_temp_c = -196

    # Build display pairs (combine first two if same pressure, adjust min temp)
    display_pairs = []
    i = 0
    while i < len(pt_pairs):
        if i == 0 and len(pt_pairs) > 1 and pt_pairs[0][1] == pt_pairs[1][1]:
            # Use material min temp instead of table min temp
            low_t = min(min_temp_c, pt_pairs[0][0])
            display_pairs.append((f"{low_t} to {pt_pairs[1][0]}", pt_pairs[0][1]))
            i += 2
        else:
            display_pairs.append((pt_pairs[i][0], pt_pairs[i][1]))  # Keep as int
            i += 1

    # Limit to 7 display pairs — covers up to 300°C for CS / SS
    # Column pairs: B:C(2), D:E(4), F:G(6), H:I(8), J:K(10), L:M(12), N:O(14)
    display_pairs = display_pairs[:7]

    # Row 8: Press., barg — Row 9: Temp., °C
    wc(8, 1, "Press., barg", left=True)
    wc(9, 1, "Temp., \u00b0C", left=True)

    pt_starts = [2, 4, 6, 8, 10, 12, 14, 16]
    max_pt = min(len(display_pairs), 8)
    for i in range(max_pt):
        c = pt_starts[i]
        mw(8, c, c + 1, display_pairs[i][1])
        mw(9, c, c + 1, display_pairs[i][0])
    for i in range(max_pt, 8):
        c = pt_starts[i]
        mw(8, c, c + 1, "")
        mw(9, c, c + 1, "")

    # Hydrotest pressure — placed right after the last P-T pair
    # Pairs use cols 2-17 (8 pairs × 2 cols), so hydrotest starts at col 18
    hydro_start = pt_starts[max_pt - 1] + 2 if max_pt > 0 else 18
    if hydro_start < 18:
        hydro_start = 18
    mw(8, hydro_start, LAST_COL, "Hydrotest Pr. (barg)")

    # Hydrotest = 1.5 × max allowable pressure at ambient
    ambient_pressure = pt_pairs[0][1] if pt_pairs else 0
    hydro_barg = round(ambient_pressure * 1.5, 1) if ambient_pressure else ""
    mw(9, hydro_start, LAST_COL, hydro_barg)

    # ══════════════════════════════════════════════════════════════
    # PIPE DATA (Rows 10-18)
    # ══════════════════════════════════════════════════════════════
    mw(10, 1, LAST_COL, "Pipe Data", bold=True)
    ws.cell(row=10, column=1).fill = fill_section
    for c in range(2, LAST_COL + 1):
        ws.cell(row=10, column=c).fill = fill_section

    # Row 11: Code
    wc(11, 1, "Code")
    pipe_std = "ASME B 36.19M" if material_type == "SS" else "ASME B 36.10M"
    mw(11, 2, LAST_COL, pipe_std)

    # Row 12: Size (in)
    wc(12, 1, "Size(in)")
    for i, ps in enumerate(pipe_sizes):
        wc(12, SC + i, float(ps["nps"]))

    # Row 13: O.D. mm
    wc(13, 1, "O.D.mm")
    for i, ps in enumerate(pipe_sizes):
        wc(13, SC + i, ps["od_mm"])

    # Row 14: Sch. (numeric schedules as numbers, text like "XS" as strings)
    wc(14, 1, "Sch.")
    for i, ps in enumerate(pipe_sizes):
        sch_val = ps["schedule"]
        try:
            sch_val = int(sch_val)
        except (ValueError, TypeError):
            pass
        wc(14, SC + i, sch_val)

    # Row 15: WT. mm
    wc(15, 1, "WT. mm")
    for i, ps in enumerate(pipe_sizes):
        wc(15, SC + i, ps["wt_mm"])

    # Row 16: TYPE — per-NPS pipe type (Seamless / LSAW / LSAW, 100% RT)
    wc(16, 1, "TYPE")
    # Group consecutive NPS sizes with the same pipe type into merged ranges
    i = 0
    while i < len(pipe_sizes):
        ptype = pipe_sizes[i]["pipe_type"]
        j = i
        while j < len(pipe_sizes) and pipe_sizes[j]["pipe_type"] == ptype:
            j += 1
        mw(16, SC + i, SC + j - 1, ptype)
        i = j

    # Row 17: MOC (different materials for seamless vs welded)
    wc(17, 1, "MOC")
    moc_smls = PIPE_MOC_SEAMLESS.get(eff_mat, f"ASTM {msr.get('material_grade', '')}")
    moc_weld = PIPE_MOC_WELDED.get(eff_mat, moc_smls)
    # Group by Seamless vs LSAW MOC (same logic as pipe type split)
    i = 0
    while i < len(pipe_sizes):
        is_seamless = pipe_sizes[i]["pipe_type"] == "Seamless"
        moc = moc_smls if is_seamless else moc_weld
        j = i
        while j < len(pipe_sizes) and (pipe_sizes[j]["pipe_type"] == "Seamless") == is_seamless:
            j += 1
        mw(17, SC + i, SC + j - 1, moc)
        i = j

    # Row 18: Ends
    wc(18, 1, "Ends")
    mw(18, 2, LAST_COL, "BE")

    # ══════════════════════════════════════════════════════════════
    # FITTINGS DATA (Rows 19-30)
    # ══════════════════════════════════════════════════════════════
    is_galv = pms_material_type.upper().replace(" ", "_") in ("CS_GALV", "CS_GALVANIZED")
    galv_suffix = ", Galvanized" if is_galv else ""

    # Forged material (small bore) / BW material (large bore)
    _forged_base = {
        "CS": "ASTM A105", "CS-LT": "ASTM A 350 Gr. LF2",
        "SS": "ASTM A 182 F316", "Alloy": "ASTM A 182 F11",
    }
    _bw_base = {
        "CS": "ASTM A 234 Gr. WPB", "CS-LT": "ASTM A 420 Gr. WPL6",
        "SS": "ASTM A 403 WP316", "Alloy": "ASTM A 234 WP11",
    }
    _olet_base = {
        "CS": "ASTM A 105", "CS-LT": "ASTM A 350 Gr. LF2",
        "SS": "ASTM A 182 F316", "Alloy": "ASTM A 182 F11",
    }
    base_mt = material_type if material_type in _forged_base else "CS"
    forged_mat = _forged_base[base_mt]
    bw_mat_fit = _bw_base[base_mt]
    olet_base_mat = _olet_base[base_mt]

    # CS GALV uses Normalized forged material
    if is_galv and forged_mat == "ASTM A105":
        forged_mat = "ASTM A 105N"
    if is_galv and olet_base_mat == "ASTM A 105":
        olet_base_mat = "ASTM A 105N"

    # Small bore boundary: NPS ≤ 1.5" = first 4 sizes (0.5, 0.75, 1, 1.5)
    SB_COUNT = 4  # indices 0-3
    sb_end_col = SC + SB_COUNT - 1   # last col of small bore (NPS 1.5")
    lb_start_col = SC + SB_COUNT     # first col of large bore (NPS 2")

    def _section_header(row, title):
        mw(row, 1, LAST_COL, title, bold=True)
        ws.cell(row=row, column=1).fill = fill_section
        for c in range(2, LAST_COL + 1):
            ws.cell(row=row, column=c).fill = fill_section

    _section_header(19, "Fittings Data")

    # Determine fittings template:
    # - CS GALV → Template B: SW small bore + BW large bore (utility)
    # - All others → Template A: All BW (process)
    use_sw_small_bore = is_galv  # Only CS GALV uses socket weld / screwed for small bore

    if use_sw_small_bore:
        # ── Template B: CS GALV (SW small bore + BW large bore) ──────

        # Row 20: TYPE
        wc(20, 1, "TYPE")
        if is_galv:
            sb_type = "Screwed (SCRD), #3000"
        else:
            sb_type = "Socket Weld (SW), #3000"
        mw(20, SC, sb_end_col, sb_type)
        mw(20, lb_start_col, LAST_COL, "Butt Weld (SCH to match pipe), Seamless")

        # Row 21: MOC
        wc(21, 1, "MOC")
        sb_moc = forged_mat + galv_suffix
        lb_moc = bw_mat_fit + ", Seamless" + galv_suffix
        mw(21, SC, sb_end_col, sb_moc)
        mw(21, lb_start_col, LAST_COL, lb_moc)

        # Rows 22-25: Elbow, Tee, Red., Cap — B16.11 (small) / B16.9 (large)
        for label, row in [("Elbow", 22), ("Tee", 23), ("Red.", 24), ("Cap", 25)]:
            wc(row, 1, label)
            mw(row, SC, sb_end_col, "ASME B 16.11")
            mw(row, lb_start_col, LAST_COL, "ASME B 16.9")

        # Row 26: Coupl (small bore only)
        wc(26, 1, "Coupl")
        mw(26, SC, sb_end_col, "ASME B 16.11")
        mw(26, lb_start_col, LAST_COL, "")

        # Row 27: Hex Hd.Plug (small bore only)
        wc(27, 1, "Hex Hd.Plug")
        mw(27, SC, sb_end_col, "Hex Head Plug, ASME B 16.11")
        mw(27, lb_start_col, LAST_COL, "")

        # Row 28: Union
        wc(28, 1, "Union")
        mw(28, SC, sb_end_col, "ASME B 16.11")
        mw(28, lb_start_col, LAST_COL, "BS 3799")

        # Row 29: Olet (MSS SP 97)
        wc(29, 1, "Olet")
        mw(29, SC, sb_end_col, "MSS SP 97")
        olet_desc_lb = f"MSS SP 97, {olet_base_mat}{galv_suffix}"
        mw(29, lb_start_col, LAST_COL, olet_desc_lb)

        # Row 30: Swage (MSS SP 95)
        wc(30, 1, "Swage")
        swage_desc = f"MSS SP 95, MOC Same as pipe{galv_suffix}"
        mw(30, 2, LAST_COL, swage_desc)

    else:
        # ── Template A: Process (All BW, no socket weld section) ─────

        # Pipe type boundary: NPS ≤ 16" = Seamless, NPS ≥ 18" = Welded
        seam_end_col = SC + next((i for i, n in enumerate(STANDARD_NPS_SIZES) if float(n) >= 18), len(STANDARD_NPS_SIZES)) - 1
        weld_start_col = seam_end_col + 1

        # Row 20: TYPE — split by Seamless / Welded (matching pipe type)
        wc(20, 1, "TYPE")
        mw(20, SC, seam_end_col, "Butt Weld (SCH to match pipe), Seamless")
        if weld_start_col <= LAST_COL:
            mw(20, weld_start_col, LAST_COL, "Butt Weld (SCH to match pipe), Welded")

        # Row 21: MOC — same BW material for all sizes
        wc(21, 1, "MOC")
        mw(21, SC, LAST_COL, bw_mat_fit)

        # Rows 22-25: Elbow, Tee, Red., Cap — all ASME B 16.9
        for label, row in [("Elbow", 22), ("Tee", 23), ("Red.", 24), ("Cap", 25)]:
            wc(row, 1, label)
            mw(row, SC, LAST_COL, "ASME B 16.9")

        # Row 26: Plug (Hex Head, ASME B 16.11 — small bore only)
        wc(26, 1, "Plug")
        mw(26, SC, sb_end_col, "Hex Head, ASME B 16.11")
        mw(26, lb_start_col, LAST_COL, "")

        # Row 27: Weldolet (MSS SP 97 — all sizes)
        wc(27, 1, "Weldolet")
        olet_mat_n = olet_base_mat
        if "A105" in olet_mat_n.replace(" ", "") or "A 105" in olet_mat_n:
            olet_mat_n = "ASTM A 105N"
        mw(27, SC, LAST_COL, f"MSS SP 97, {olet_mat_n}")

        # Rows 28-30: not used for process template (no Coupling/Union/Swage)
        # Leave empty — these rows will be hidden
        for row in [28, 29, 30]:
            for c in range(1, LAST_COL + 1):
                ws.cell(row=row, column=c).value = None
            ws.row_dimensions[row].height = 0  # Hide blank rows

    # ══════════════════════════════════════════════════════════════
    # FLANGE (Rows 31-35)
    # ══════════════════════════════════════════════════════════════
    _section_header(31, "Flange")

    fl_moc = FLANGE_MOC.get(base_mt, fl.get("material", "ASTM A 105"))
    if is_galv:
        fl_moc = fl_moc + " Galvanized" if "Galvanized" not in fl_moc else fl_moc

    if use_sw_small_bore:
        # CS GALV: SW/SCRD for small bore, WN for large bore
        wc(32, 1, "TYPE")
        if is_galv:
            fl_sb_type = "Screwed (SCRD)"
        else:
            fl_sb_type = "Socket Weld (SW)"
        mw(32, SC, sb_end_col, fl_sb_type)
        mw(32, lb_start_col, LAST_COL, "WN")
    else:
        # Process: WN for all sizes, with standard reference
        wc(32, 1, "MOC")
        fl_moc_n = fl_moc
        if "A105" in fl_moc_n.replace(" ", "") or "A 105" in fl_moc_n:
            fl_moc_n = "ASTM A 105N"
        mw(32, 2, LAST_COL, fl_moc_n)

    # Row 33: MOC (Flange) / FACE
    if use_sw_small_bore:
        wc(33, 1, "MOC")
        mw(33, 2, LAST_COL, fl_moc)
    else:
        face_desc = f"{piping_class}# {face_type}"
        if face_type == "RF":
            face_desc += ", Serrated Finish"
        wc(33, 1, "FACE")
        mw(33, 2, LAST_COL, face_desc)

    # Row 34: FACE / TYPE
    if use_sw_small_bore:
        face_desc = f"{piping_class}# {face_type}"
        if face_type == "RF":
            face_desc += ", Serrated Finish"
        wc(34, 1, "FACE")
        mw(34, 2, LAST_COL, face_desc)
    else:
        wc(34, 1, "TYPE")
        mw(34, 2, LAST_COL, f"Weld Neck, ASME B 16.5/ 16.47A, Butt Welding ends as per ASME B 16.25")

    # Row 35: STD
    wc(35, 1, "STD")
    mw(35, 2, LAST_COL, "ASME B 16.5")

    # ══════════════════════════════════════════════════════════════
    # SPECTACLE BLIND / SPACER BLINDS (Rows 36-38)
    # ══════════════════════════════════════════════════════════════
    _section_header(36, "Spectacle Blind/Spacer Blinds")

    wc(37, 1, "MOC")
    spec_blind_moc = fl_moc
    if "A105" in spec_blind_moc.replace(" ", "") or "A 105" in spec_blind_moc:
        if "N" not in spec_blind_moc and "Galvanized" not in spec_blind_moc:
            spec_blind_moc = "ASTM A 105N"
        elif "Galvanized" in spec_blind_moc:
            spec_blind_moc = "ASTM A 105N Galvanized"
    mw(37, 2, LAST_COL, spec_blind_moc)

    wc(38, 1, "Spectacle")
    mw(38, 2, 12, "ASME B 16.48")
    mw(38, 13, LAST_COL, "Spacer and blind as per ASME B 16.48")

    # ══════════════════════════════════════════════════════════════
    # BOLTS / NUTS / GASKETS (Rows 39-42)
    # ══════════════════════════════════════════════════════════════
    _section_header(39, "Bolts/ Nuts/ Gaskets")

    # For CS GALV bolting, use the CS key (base material type)
    _bolt_key = (str(base_mt), bool(is_nace))
    bolt_mat = BOLT_MATERIAL.get(_bolt_key, "ASTM A 193 Gr. B7")
    nut_mat = NUT_MATERIAL.get(_bolt_key, "ASTM A 194 Gr. 2H")
    gasket_desc = gk.get("description") or GASKET_DESC.get(face_type, GASKET_DESC["RF"])

    # CS GALV gasket override: neoprene/EPDM rubber flat ring
    if is_galv:
        gasket_desc = "3mm thick flat ring of neoprene/ EPDM rubber as ASME B 16.21"

    wc(40, 1, "Stud Bolt")
    mw(40, 2, LAST_COL, bolt_mat)

    wc(41, 1, "Hex Nuts")
    mw(41, 2, LAST_COL, nut_mat)

    wc(42, 1, "Gasket")
    mw(42, 2, LAST_COL, gasket_desc)

    # ══════════════════════════════════════════════════════════════
    # VALVES (Rows 43-52)
    # ══════════════════════════════════════════════════════════════
    _section_header(43, "Valves")

    try:
        pc_int = int(piping_class)
    except (TypeError, ValueError):
        pc_int = 150

    wc(44, 1, "Rating")
    mw(44, 2, LAST_COL, f"{piping_class}#, {face_type}")

    # ── Generate VDS codes using recommendation engine ─────────────
    vds_rec = recommend_vds_for_class(
        spec_code=pms_code, material_type=material_type,
        pressure_class=pc_int, is_nace=is_nace,
        pms_material_type=pms_material_type,
    )

    # Small/large bore column split (same as fittings)
    VALVE_SMALL_SPLIT = SB_COUNT  # 4 sizes: 0.5, 0.75, 1, 1.5
    v_sb_end = SC + VALVE_SMALL_SPLIT - 1
    v_lb_start = SC + VALVE_SMALL_SPLIT

    # Index VDS by valve type for lookup
    sb_vds = {v["valve_type"]: v["vds_code"] for v in vds_rec["small_bore"]}
    lb_vds = {v["valve_type"]: v["vds_code"] for v in vds_rec["large_bore"]}

    # Butterfly start col: NPS ≥ 6" = index 7 in STANDARD_NPS_SIZES → col SC+7
    bf_start_idx = next((i for i, n in enumerate(STANDARD_NPS_SIZES) if float(n) >= 6), 7)
    bf_start_col = SC + bf_start_idx

    # Check valve: Piston for small, Swing + Dual Plate for large
    # Dual Plate typically ≥ 6" alongside Swing
    check_sb = sb_vds.get("Check", "")
    check_swing = lb_vds.get("Check", "")
    check_dp = lb_vds.get("Check (Dual Plate)", "")

    valve_rows = [
        (45, "Ball"),
        (46, "Gate"),
        (47, "Globe"),
        (48, "Check"),
        (49, "Butterfly"),
        (50, "Needle"),
    ]

    for row, vtype in valve_rows:
        wc(row, 1, vtype)
        sb_tag = sb_vds.get(vtype, "")
        lb_tag = lb_vds.get(vtype, "")

        if vtype == "Check":
            # Small bore: Piston check
            mw(row, SC, v_sb_end, check_sb)
            # Large bore: Swing check + Dual Plate
            if check_dp:
                mw(row, v_lb_start, LAST_COL, f"{check_swing}, {check_dp}")
            else:
                mw(row, v_lb_start, LAST_COL, check_swing)
        elif vtype == "Butterfly":
            # Butterfly only for NPS ≥ 6"
            mw(row, SC, bf_start_col - 1, "")
            mw(row, bf_start_col, LAST_COL, lb_tag)
        elif vtype == "Needle":
            # Needle only for small bore
            mw(row, SC, v_sb_end, sb_tag)
            mw(row, v_lb_start, LAST_COL, "")
        elif sb_tag == lb_tag:
            mw(row, SC, LAST_COL, sb_tag)
        else:
            mw(row, SC, v_sb_end, sb_tag)
            mw(row, v_lb_start, LAST_COL, lb_tag)

    # Empty rows 51-52 for additional valves (DBB, etc.)
    for r in [51, 52]:
        wc(r, 1, "")
        mw(r, 2, LAST_COL, "")

    # ══════════════════════════════════════════════════════════════
    # NOTES (Rows 53+)
    # ══════════════════════════════════════════════════════════════
    _section_header(53, "Notes")

    notes = [
        f"1. PMS to be read in conjunction with Piping Design Basis, Doc No. {meta.get('doc_number', '50501-SPE- 80000-PP-RL-0001')} and Valve Material Specification.",
        "2. Weld Joint Factor for welded pipe shall be as per ASME B 31.3.",
        "3. Welded fittings shall be 100% radiographed.",
        "4. Two jackscrew, 180 degree apart shall be provided in one of the flanges for all orifice flange and specified spectacle blind assemblies.",
        "5. Spectacle blinds and spacer sizes and rating that are not available in ASME B 16.48 shall be as per manuf. standard. Design shall be submitted to Company for review and approval.",
        "6. Maximum temperature limit for all Soft Seat Ball Valve shall be 200\u00b0C.",
        "7. Wafer check valve to be avoided, unless the available space constraint does not allow normal check valve. ",
    ]

    for i, note in enumerate(notes):
        mw(54 + i, 1, LAST_COL, note, left=True)

    last_row = 54 + len(notes)

    # ══════════════════════════════════════════════════════════════
    # ROW HEIGHTS
    # ══════════════════════════════════════════════════════════════
    for r in range(1, 7):
        ws.row_dimensions[r].height = 16
    # Section headers
    for r in [7, 10, 19, 31, 36, 39, 43, 53]:
        ws.row_dimensions[r].height = 20
    for r in range(12, 16):
        ws.row_dimensions[r].height = 15
    # Fittings (20-30), Flange (32-35), Bolts (40-42), Valves (44-52)
    for r in list(range(20, 31)) + list(range(32, 36)) + list(range(37, 43)) + list(range(44, 53)):
        ws.row_dimensions[r].height = 28
    for r in range(54, last_row):
        ws.row_dimensions[r].height = 28

    # ══════════════════════════════════════════════════════════════
    # FINAL BORDER PASS — fill empty cells + medium outer edges
    # ══════════════════════════════════════════════════════════════
    for r in range(1, last_row):
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.border or cell.border == Border():
                cell.border = tb
            if not cell.font or cell.font == Font():
                cell.font = font_n
                cell.alignment = ac

    for r in range(1, last_row):
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            existing = cell.border
            top = ms if r == 1 else (existing.top if existing else ts)
            bottom = ms if r == last_row - 1 else (existing.bottom if existing else ts)
            left = ms if c == 1 else (existing.left if existing else ts)
            right = ms if c == LAST_COL else (existing.right if existing else ts)
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # ── Page setup ─────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3

    # ── Save ───────────────────────────────────────────────────────
    sc_code = sc.get("spec_code", "PMS")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"PMS_{sc_code}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n  Excel PMS saved to: {filepath}")
    return filepath


def generate_pms(pms_data, output_dir=None):
    """Generate complete PMS document (console + Excel)."""
    generate_pms_console(pms_data)
    if output_dir:
        return generate_pms_excel(pms_data, output_dir)
    return None
